# -*- coding: utf-8 -*-
'''按照每一行读取文件gdb.log的内容'''
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment


with open('192.168.0.14_2024-11-10_16-41-26.log', 'r') as file:
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    addrMap = {}
    info = {}
    ws.title = "Memory Allocations"
    # 写入表头
    ws.append(["Thread", "Size", "Stack Trace"])
    flag = False
    stack_trace = ""
    size = 0
    addr = None
    addnum = 0
    delnum = 0
    delsize = 0
    thread = ""
    for line in file:
        # Process each line here
        # print(line)
        if "Switching to LWP" in line:
            thread = line.split("LWP ")[1].split(']')[0]
            # print(thread)
        if "Breakpoint" in line:
            # print(line)
            if flag:
                info["stack"] = stack_trace
                addrMap[addr] = info.copy()
                # ws.append([size, stack_trace])
            flag = False
            stack_trace = ""
            info.clear()
            size = 0
            addr = ""
            if "Breakpoint 4, zl_malloc" in line:
                addnum += 1
                flag = True
                # 获取行line中size的值
                size = line.split("sizereal=")[1].split(')')[0]
                addr = line.split("ptr=")[1].split(',')[0]
                info["size"] = size
                info["thread"] = thread
            elif "Breakpoint 5, zl_free" in line:
                delnum += 1
                addr = line.split("ptr=")[1].split(',')[0]
                if addr in addrMap:
                    delsize += 1
                    addrMap.pop(addr)
        elif flag and '#' in line:
            if len(stack_trace) > 0:
                stack_trace += '\r\n'
            stack_trace += line.split(']')[1]
    # print(addrMap)
    if flag:
        info["stack"] = stack_trace
        addrMap[addr] = info
    print(len(addrMap))
    print(addnum)
    print(delnum)
    print(delsize)

    threadmap = {}
    
    sum = 0
    # 将addrMap表里的数据写入ws中
    for addr, info in addrMap.items():
        # print(addr)
        # print(info)
        # ws.append([info["size"], info["stack"]])
        
        size = int(info["size"])
        sum += size
        threadmap[info["thread"]] = threadmap.get(info["thread"], 0) + 1
        # row = [info["thread"], info["size"], info["stack"]]
        # ws.append(row)
        # # 获取最后一行的单元格
        # cell = ws.cell(row=ws.max_row, column=3)
        # # 设置单元格的对齐方式，启用换行
        # cell.alignment = Alignment(wrap_text=True)
    # 保存 Excel 文件
    # wb.save("memory_allocations-1110.xlsx")
    print(sum)
    #json格式化打印threadmap的信息
    print(json.dumps(threadmap, indent=4))



